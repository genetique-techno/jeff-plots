#! /usr/bin/env python3

import re
import plotly.express as px
from openpyxl import load_workbook
from os import mkdir, scandir
from pathlib import Path
import datetime

#
# --- User Params ---
#
# file: Name of the file to be processed. Must be located in the `data` directory.
# sample_date_row: 1-indexed row number that contains dates for the x-axis.
# date_validity_check_regex: Regex rule for checking that the dates are valid. Is used to determine the end of the useful columns on a sheet.
# analyte_column: 1-indexed column number that contains the analyte names.
# units_column: 1-indexed column number that contains the units for each analyte.
# min_row: 1-indexed row where measurement data starts.
# max_row: 1-indexed row where measurement data ends. Should be used for testing only, may cause side effects. `None` directs the scipt to find the max.
# min_column: 1-indexed column where measurement data starts.
# max_column: 1-indexed column where sample dates. Should be used for testing only, may cause side effects. `None` directs the scipt to find the max.
# exclude_sheets: Array of sheet names to exclude from processing.

file = "new.xlsm"
sample_date_row = 2
date_validity_check_regex = "^[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}$"
analyte_column = 1
units_column = 4
min_row = 4
max_row = None
min_column = 5
max_column = None
exclude_sheets = ["CL-1", "CL-2", "CL-3", "CL-4", "CL-5", "Sheet3", "Sheet4"]

#
# --- Plot Styling Params ---
#
# x_label: Label to apply to the x-axis.
# series_label: Label to apply to the species legend
# width: Width (px) of output images.
# height: Height (px) of output images.

x_label = "Sample Date"
series_label = "Wells"
width = 700
height = 500

#
# --- Helper Methods ---
#
def remove_units(value):
  """Removes the units from a value that is a string. Otherwise returns the value untouched."""
  if type(value) == str:
    return value.split()[0]
  else:
    return value

def convert_no_sample_to_none(value):
  """Converts a NS value to a None."""
  if value == "NS":
    return None
  else:
    return value

def convert_non_detect_to_zero(value):
  """Converts value below measurement threshold to zero. Otherwise returns the value untouched."""
  if type(value) == str and value.startswith("<"):
    return 0
  else:
    return value

def convert_to_float(value):
  """Converts a value to a float."""
  if value != None:
    return float(value)
  else:
    return value

def get_values_from_cells(cells):
  """Takes a list of cells and returns the values of the cells."""
  return [ c.value for c in cells ]

def merge_dict_data(dict1, dict2):
  """Merge dictionaries that contain arrays values and concat the arrays of keys found in both. Ignores the key `y_label`."""
  result = {**dict1, **dict2}
  for key, value in result.items():
    if key != "y_label" and key in dict1 and key in dict2:
      result[key] = dict1[key] + value
    else:
      result[key] = value
  return result

def convert_to_datetime(dateStr):
  """Takes a mm/dd/yyyy string and converts it to a datetime.date python object."""
  if type(dateStr) == datetime.datetime:
    return dateStr.isoformat()
  else:
    arr = dateStr.split("/")
    arr = [ int(x) for x in arr ]
    try:
      month, day, year = arr
    except:
      print("Broken dateStr object: " + dateStr)

    return datetime.datetime(year, month, day, 0, 0).isoformat()

def get_max_column(sheet):
  """Takes a sheet. Takes the sample date row and searches for the first `None` value in the cells and returns the cell.column."""
  global min_column
  global max_column
  global sample_date_row
  if max_column == None:
    for cell_column in sheet.iter_cols(min_col = min_column, max_col = 1000):
      if max_column == None:
        if cell_column[sample_date_row].value == None:
          max_column = cell_column[sample_date_row].column - 1
  else:
    return max_column

def get_max_row(sheet):
  """Takes a sheet. Takes the min_row and searches for the first `None` value in the cells and returns the cell.row."""
  global min_row
  global max_row
  global analyte_column
  if max_row == None:
    for cell_row in sheet.iter_rows(min_row, max_row = 1000):
      if max_row == None:
        if cell_row[analyte_column - 1].value == None:
          max_row = cell_row[analyte_column - 1].row - 1
  else:
    return max_row

#
# --- Data Extraction Methods ---
#
def get_date_cells(sheet, sample_date_row):
  """Takes the worksheet, uses sample_date_row & min_column to return just the date values."""
  global max_column
  for a_row in sheet.iter_rows(min_row = sample_date_row, max_row = sample_date_row):
    row = a_row
  cells = [ c for c in row if c.column >= min_column ]
  cells = [ c for c in cells if c.column <= max_column ]
  return cells

def get_cell_value_from_row(row, column_num):
  """Takes a openpyxl Row instance, and a 0-indexed column number, returns the value of that cell."""
  return row[column_num - 1].value

def get_analyte_cells(row, min_column, max_column):
  """Returns the data values for the row, using min_column and max_column to narrow the dataset."""
  return [ c for c in row if c.column >= min_column and c.column <= max_column ]

#
# --- Combinatorial Methods ---
#
def process_sheet(sheet):
  """Given a Sheet instance, returns a dict of all analytes with data_frame dicts for each."""
  get_max_column(sheet)
  get_max_row(sheet)
  date_cells = get_date_cells(sheet, sample_date_row)
  date_values = get_values_from_cells(date_cells)
  for i in range(len(date_values)):
    date_values[i] = convert_to_datetime(date_values[i])
  sheet_result = {}
  for row in sheet.iter_rows(min_row, max_row):
    analyte_values = get_values_from_cells(get_analyte_cells(row, min_column, max_column))
    analyte_values = list(map(remove_units, analyte_values))
    analyte_values = list(map(convert_no_sample_to_none, analyte_values))
    analyte_values = list(map(convert_non_detect_to_zero, analyte_values))
    analyte_values = list(map(convert_to_float, analyte_values))
    site_labels = [sheet.title for i in range(len(analyte_values))]
    analyte_name = get_cell_value_from_row(row, analyte_column)
    units = get_cell_value_from_row(row, units_column)
    # error check needed: len(date_values) == len(analyte_values) == len(site_labels) or bad things happen
    sheet_result.update({ analyte_name: { x_label: date_values, "y_label": units, "y_values": analyte_values, series_label: site_labels } })
  return sheet_result

def process_workbook(book, sheetnames):
  """Takes a workbook and a list of sheet names. Returns a data dict with a key for each analyte collected from all relevant sheets."""
  book_result = {}
  for sheetname in sheetnames:
    sheet = book[sheetname]
    sheet_result = process_sheet(sheet)
    for analyte, analyte_data in sheet_result.items():
      if analyte in book_result:
        book_result[analyte] = merge_dict_data(book_result[analyte], analyte_data)
      else:
        book_result[analyte] = analyte_data
  return book_result

def make_plot(analyte, data):
  """Takes data, generates a scatter plot and saves the image to a file."""
  plot = px.scatter(data_frame = data, x = x_label, y = "y_values", labels = { "y_values": data["y_label"] }, color = series_label, title = analyte, width = width, height = height)
  bytes = plot.to_image(format = "png")
  filename = analyte + ".png"
  outfile = open(Path("output/") / filename, "wb")
  outfile.write(bytes)
  outfile.close()
  print(filename)
  return None

#
# --- Main Program Flow
#
print("--- Starting ---")
# Ensure the output directory exists.
try:
  scandir("output")
except:
  mkdir("output")
# Acquire the workbook.
wb = load_workbook(Path("data/") / file)
# Find only valid sheet names from the workbook.
sheetnames = [ item for item in wb.sheetnames if item not in exclude_sheets ]
# Compile a RegExp object for finding date values.
date_re = re.compile(date_validity_check_regex)
# Process the workbook into a dictionary containing all analyte data.
results = process_workbook(wb, sheetnames)
# Make plots for every analyte.
[ make_plot(analyte, data) for analyte, data in results.items() ]
print("--- Finished ---")
